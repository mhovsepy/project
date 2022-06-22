from selenium import webdriver
import openpyxl 
from bs4 import BeautifulSoup as bs
import time
wb = openpyxl.Workbook() 
sheet = wb.active 

driver = webdriver.Firefox(executable_path=r'C:\Users\Home\Desktop\Lyow\Data_Analytics\geckodriver\geckodriver.exe')
url = 'https://www.whoscored.com/Statistics'
time.sleep(5)
driver.get(url)
time.sleep(5)
soup = bs(driver.page_source, 'lxml')
result = soup.find('table', {'id':'top-team-stats-summary-grid'})
headers = result.find_all('th')
sheet.cell(row = 1, column = 1).value = 'Teamid'
sheet.cell(row = 1, column = 2).value = 'Teamname'
sheet.cell(row = 1, column = 3).value = 'Tournament'
sheet.cell(row = 1, column = 4).value = 'Goals'
sheet.cell(row = 1, column = 5).value = 'Shots pg'
sheet.cell(row = 1, column = 6).value = 'Yellow'
sheet.cell(row = 1, column = 7).value = 'Red'
sheet.cell(row = 1, column = 8).value = 'Possession%'
sheet.cell(row = 1, column = 9).value = 'Pass%'
sheet.cell(row = 1, column = 10).value = 'AerialsWon'
sheet.cell(row = 1, column = 11).value = 'Rating'

def pages(page):
    time.sleep(3)
    nexts = driver.find_elements_by_id('next')
    driver.execute_script("document.body.style['-webkit-transform'] = \"scale(0.5)\";")
    soup = bs(driver.page_source, 'lxml')
    result = soup.find('table', {'id':'top-team-stats-summary-grid'})
    rows = result.find_all('td', class_ = 'col12-lg-2 col12-m-3 col12-s-4 col12-xs-5 grid-abs overflow-text')
    for i in range(20):
        try:
            sheet.cell(row = page*20 + i + 2, column = 1).value = rows[i].text.split('. ')[0]
            sheet.cell(row = page*20 + i + 2, column = 2).value = rows[i].text.split('. ')[1]
        except:
            continue
  
    tbody = result.find('tbody', {'id':'top-team-stats-summary-content'})
    tr = tbody.find_all('tr')
    for j in range(len(tr)):
        td1 = tr[j].find_all('td')
        for i in range(1,4):
            sheet.cell(row = page*20 + j + 2, column = i + 2).value = td1[i].text
        sheet.cell(row = page*20 + j + 2, column = 6).value = td1[4].find_all('span')[0].text
        sheet.cell(row = page*20 + j + 2, column = 7).value = td1[4].find_all('span')[1].text
        for i in range(4):
            sheet.cell(row = page*20 + j + 2, column = i + 8).value = td1[i+5].text
    time.sleep(1)
    wb.save(r"C:\Users\Home\Desktop\Lyow\Data_Analytics\Python\0Project\football\teams\teamstable.csv")
    nexts[0].click()
for i in range(5):
    pages(i)
wb.save(r"C:\Users\Home\Desktop\Lyow\Data_Analytics\Python\0Project\football\teams\teamstable.csv")
driver.quit()