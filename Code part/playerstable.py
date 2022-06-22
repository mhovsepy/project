from selenium import webdriver
import openpyxl 
from bs4 import BeautifulSoup as bs
import time
wb = openpyxl.Workbook() 
sheet = wb.active 

driver = webdriver.Firefox(executable_path=r'C:\Users\Home\Desktop\Lyow\Data_Analytics\geckodriver\geckodriver.exe')
url = 'https://www.whoscored.com/Statistics'
time.sleep(5)
driver.get(url)
time.sleep(5)
soup = bs(driver.page_source, 'lxml')
result = soup.find('table', {'id':'top-player-stats-summary-grid'})
headers = result.find_all('th')
sheet.cell(row = 1, column = 3).value = 'Team'
sheet.cell(row = 1, column = 4).value = 'Age'
for i in headers:
    if headers.index(i) + 2 >=4:
        sheet.cell(row = 1, column = headers.index(i) + 3).value = i.text
    elif headers.index(i) + 2 <=2:
        sheet.cell(row = 1, column = headers.index(i) + 1).value = i.text
def pages(page):
    time.sleep(3)
    nexts = driver.find_elements_by_id('next')
    driver.execute_script("document.body.style['-webkit-transform'] = \"scale(0.5)\";")
    soup = bs(driver.page_source, 'lxml')
    result = soup.find('table', {'id':'top-player-stats-summary-grid'})
    rows = result.find_all('td', class_ = 'col12-lg-2 col12-m-3 col12-s-4 col12-xs-5 grid-abs overflow-text')
    teamnames = result.find_all('span',class_ = 'team-name')
    ages = result.find_all('span',class_ = 'player-meta-data')
    for i in range(10):
        sheet.cell(row = page*10 + i + 2, column = 3).value = teamnames[2*i].text
    for i in range(10):
        sheet.cell(row = page*10 + i + 2, column = 4).value = ages[4*i].text
        
    tbody = result.find('tbody', {'id':'player-table-statistics-body'})
    for row in rows:
        rownumber = row.find('div', class_ = 'table-ranking')
        name = row.find('span', {'class':['iconize','iconize-icon-left']})
        sheet.cell(row = page*10 + rows.index(row) + 2, column = 1).value = int(rownumber.text)
        sheet.cell(row = page*10 + rows.index(row) + 2, column = 2).value = name.text
    tbody = result.find('tbody', {'id':'player-table-statistics-body'})
    tr = tbody.find_all('tr')
    for j in range(len(tr)):
        td1 = tr[j].find_all('td')
        for i in range(2,len(td1)):
            sheet.cell(row = page*10 + j + 2, column = i + 3).value = td1[i].text
    time.sleep(1)
    wb.save(r"C:\Users\Home\Desktop\Lyow\Data_Analytics\Python\0Project\football\players\playerstable.csv")
    nexts[1].click()
for i in range(161):
    pages(i)
wb.save(r"C:\Users\Home\Desktop\Lyow\Data_Analytics\Python\0Project\football\players\playerstable.csv")
driver.quit()